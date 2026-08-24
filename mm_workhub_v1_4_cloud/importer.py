from __future__ import annotations

import re
from datetime import datetime, time, timedelta
from io import BytesIO
from typing import BinaryIO

import openpyxl

from db import connect, init_db

STAGES = [
    "Analysing/Designing",
    "Functional Spec",
    "Sr Review",
    "ABAP Estimation",
    "ADL APPROVAL DEV",
    "ABAP DEV",
    "ABAP Fixing",
    "Unit Testing",
    "UAT",
    "External Approval",
    "BASIS",
    "FMP",
]

DATE_PATTERNS = [
    re.compile(r"\[(\d{1,2})[-./](\d{1,2})[-./](\d{4})\]\s*$"),
    re.compile(r"\[(\d{1,2})[-./](\d{1,2})[-./](\d{2})\]\s*$"),
]


def _iso_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value)


def _iso_time(value):
    if value is None:
        return None
    if isinstance(value, (datetime, time)):
        return value.strftime("%H:%M:%S")
    return str(value)


def parse_history(text: str | None):
    if not text:
        return []
    rows = []
    for idx, raw in enumerate(str(text).splitlines(), start=1):
        line = raw.strip().lstrip("-•").strip()
        if not line:
            continue
        update_date = None
        clean = line
        for pattern in DATE_PATTERNS:
            m = pattern.search(line)
            if m:
                d, mth, y = map(int, m.groups())
                if y < 100:
                    y += 2000
                try:
                    update_date = datetime(y, mth, d).date().isoformat()
                except ValueError:
                    update_date = None
                clean = line[:m.start()].strip()
                break
        rows.append({"update_text": clean or line, "update_date": update_date, "source_order": idx})
    return rows


def _extract_reference(activity: str):
    m = re.match(r"\[([^\]]+)\]", activity.strip())
    return m.group(1).strip() if m else None


def _load_workbook(file_obj):
    if isinstance(file_obj, (str, bytes)):
        return openpyxl.load_workbook(file_obj, data_only=False, keep_vba=True)
    raw = file_obj.read()
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)
    return openpyxl.load_workbook(BytesIO(raw), data_only=False, keep_vba=True)


def import_workbook(file_obj, reset: bool = True):
    init_db()
    wb = _load_workbook(file_obj)
    stats = {"tickets": 0, "updates": 0, "stages": 0, "timesheets": 0}

    with connect() as conn:
        if reset:
            conn.executescript("""
            DELETE FROM ticket_updates;
            DELETE FROM ticket_stages;
            DELETE FROM timesheets;
            DELETE FROM tickets;
            """)

        ticket_id_by_ams = {}

        # Main ticket master
        ws = wb["Main"]
        for r in range(11, ws.max_row + 1):
            ams = ws.cell(r, 2).value
            if not ams:
                continue
            jira = ws.cell(r, 3).value
            description = ws.cell(r, 4).value
            created = _iso_date(ws.cell(r, 5).value)
            priority = ws.cell(r, 6).value
            pic = ws.cell(r, 7).value
            history = ws.cell(r, 8).value
            status = ws.cell(r, 9).value
            obj = ws.cell(r, 10).value
            action = ws.cell(r, 11).value
            transport = ws.cell(r, 12).value
            last_date = _iso_date(ws.cell(r, 13).value)
            last_time = _iso_time(ws.cell(r, 14).value)
            ticket_type = ws.cell(r, 15).value
            mandays = ws.cell(r, 16).value
            last_checked = ws.cell(r, 17).value
            if not isinstance(mandays, (int, float)):
                mandays = None
            if hasattr(last_checked, "text"):
                last_checked = None
            elif last_checked is not None and str(last_checked).startswith("="):
                last_checked = None

            cur = conn.execute("""
                INSERT INTO tickets (
                    ams_ticket,jira_ticket,description,date_created,priority,pic,status,
                    object_status,action_status,latest_transport,last_update_date,last_update_time,
                    ticket_type,mandays_chargeable,last_checked
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(ams_ticket) DO UPDATE SET
                    jira_ticket=excluded.jira_ticket, description=excluded.description,
                    date_created=excluded.date_created, priority=excluded.priority, pic=excluded.pic,
                    status=excluded.status, object_status=excluded.object_status,
                    action_status=excluded.action_status, latest_transport=excluded.latest_transport,
                    last_update_date=excluded.last_update_date, last_update_time=excluded.last_update_time,
                    ticket_type=excluded.ticket_type, mandays_chargeable=excluded.mandays_chargeable,
                    updated_at=CURRENT_TIMESTAMP
            """, (
                str(ams).strip(), str(jira) if jira is not None else None, description, created,
                priority, pic, status, obj, action, transport, last_date, last_time,
                ticket_type, mandays, last_checked
            ))
            row = conn.execute("SELECT id FROM tickets WHERE ams_ticket=?", (str(ams).strip(),)).fetchone()
            ticket_id = row[0]
            ticket_id_by_ams[str(ams).strip()] = ticket_id
            stats["tickets"] += 1

            for item in parse_history(history):
                conn.execute(
                    "INSERT INTO ticket_updates(ticket_id,update_date,update_text,source_order) VALUES (?,?,?,?)",
                    (ticket_id, item["update_date"], item["update_text"], item["source_order"]),
                )
                stats["updates"] += 1

        # Stage tracker. Resolve descriptions/jira from Main, never from formulas.
        if "Ticket tracker" in wb.sheetnames:
            ws = wb["Ticket tracker"]
            for r in range(8, ws.max_row + 1):
                ams = ws.cell(r, 2).value
                if not ams:
                    continue
                ams = str(ams).strip()
                ticket_id = ticket_id_by_ams.get(ams)
                if ticket_id is None:
                    # rare tracker-only record
                    conn.execute(
                        "INSERT INTO tickets(ams_ticket) VALUES (?) ON CONFLICT(ams_ticket) DO NOTHING",
                        (ams,),
                    )
                    ticket_id = conn.execute("SELECT id FROM tickets WHERE ams_ticket=?", (ams,)).fetchone()[0]
                    ticket_id_by_ams[ams] = ticket_id
                for i, stage in enumerate(STAGES, start=1):
                    # tracker columns F:Q = 6..17
                    raw = ws.cell(r, 5 + i).value
                    completed = 1 if raw is True or str(raw).strip().upper() == "TRUE" else 0
                    conn.execute(
                        """
                        INSERT INTO ticket_stages(ticket_id,stage_name,stage_order,completed)
                        VALUES (?,?,?,?)
                        ON CONFLICT(ticket_id,stage_name) DO UPDATE SET
                            stage_order=excluded.stage_order,
                            completed=excluded.completed
                        """,
                        (ticket_id, stage, i, completed),
                    )
                    stats["stages"] += 1

        # Timesheet matrix: merge consecutive same-activity hourly cells by date.
        if "Date" in wb.sheetnames:
            ws = wb["Date"]
            for c in range(2, ws.max_column + 1):
                work_date = ws.cell(1, c).value
                if not work_date:
                    continue
                date_iso = _iso_date(work_date)
                slots = []
                for r in range(2, ws.max_row + 1):
                    slot_time = ws.cell(r, 1).value
                    activity = ws.cell(r, c).value
                    if slot_time and activity:
                        slots.append((slot_time, str(activity).strip()))
                i = 0
                while i < len(slots):
                    start, activity = slots[i]
                    end = (datetime.combine(datetime.today(), start) + timedelta(hours=1)).time()
                    j = i + 1
                    while j < len(slots) and slots[j][1] == activity:
                        prev_t = slots[j - 1][0]
                        expected = (datetime.combine(datetime.today(), prev_t) + timedelta(hours=1)).time()
                        if slots[j][0] != expected:
                            break
                        end = (datetime.combine(datetime.today(), slots[j][0]) + timedelta(hours=1)).time()
                        j += 1
                    start_dt = datetime.combine(datetime.today(), start)
                    end_dt = datetime.combine(datetime.today(), end)
                    hours = (end_dt - start_dt).total_seconds() / 3600
                    if hours <= 0:
                        hours += 24
                    reference = _extract_reference(activity)
                    category = "Leave" if activity.upper().strip() in {"EL", "AL", "MC", "LEAVE"} else "Work"
                    conn.execute(
                        "INSERT INTO timesheets(work_date,start_time,end_time,activity,reference,hours,category) VALUES (?,?,?,?,?,?,?)",
                        (date_iso, start.strftime("%H:%M"), end.strftime("%H:%M"), activity, reference, hours, category),
                    )
                    stats["timesheets"] += 1
                    i = j

    return stats
